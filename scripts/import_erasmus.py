"""Import Erasmus_Mundus_GIS_Catalogue.xlsx into ApplyDesk (idempotent by slug)."""

from __future__ import annotations

import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

EXCEL = Path(r"D:\Upwork\00 - Agent\Erasmus Mundus\Erasmus_Mundus_GIS_Catalogue.xlsx")
ROOT = Path(__file__).resolve().parents[1]
ENV_FILES = [ROOT / ".env.local", ROOT / ".env.import"]
CYCLE_LABEL = "2027 intake"

# Parsed ISO dates from the workbook text (Pakistan / Partner Country).
# Original wording is stored on extra / extra_dates so nothing is lost.
DATES = {
    "GEM": {
        "opens": "2026-10-01",
        "scholarship": "2027-01-15",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "CDE": {
        "opens": "2026-11-01",
        "scholarship": "2027-01-18",
        "starts": "2027-10-01",
        "status": "confirmed",
        "scholarships": True,
    },
    "GeoTech": {
        "opens": "2026-11-01",
        "scholarship": "2027-01-15",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "WAVES": {
        "opens": "2026-10-01",
        "scholarship": None,
        "starts": "2027-09-01",
        "status": "tba",
        "scholarships": True,
    },
    "GroundwatCH": {
        "opens": "2026-10-01",
        "scholarship": "2027-01-05",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "FloodRISK": {
        "opens": "2026-10-01",
        "scholarship": "2027-01-05",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "Cartography": {
        "opens": "2027-01-01",
        "scholarship": None,
        "starts": "2027-10-01",
        "status": "expected",
        "scholarships": None,
    },
    "MUrCS+": {
        "opens": None,
        "scholarship": "2027-01-15",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "SUFONAMA": {
        "opens": "2026-11-01",
        "scholarship": "2027-01-15",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "MSc EF": {
        "opens": "2026-12-02",
        "scholarship": "2027-01-13",
        "starts": "2027-08-01",
        "status": "confirmed",
        "scholarships": True,
    },
    "MESPOM": {
        "opens": "2026-11-01",
        "scholarship": "2027-01-07",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "QuBiD": {
        "opens": None,
        "scholarship": None,
        "starts": "2027-09-01",
        "status": "tba",
        "scholarships": True,
    },
    "BIOVERTE": {
        "opens": None,
        "scholarship": None,
        "starts": "2027-09-01",
        "status": "tba",
        "scholarships": True,
    },
    "sosMER": {
        "opens": None,
        "scholarship": None,
        "starts": "2027-09-01",
        "status": "tba",
        "scholarships": True,
    },
    "MERGED": {
        "opens": "2026-11-15",
        "scholarship": "2027-01-15",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": True,
    },
    "IMETE": {
        "opens": None,
        "scholarship": None,
        "starts": None,
        "status": "tba",
        "scholarships": False,
    },
    "COASTHazar": {
        "opens": None,
        "scholarship": None,
        "starts": None,
        "status": "tba",
        "scholarships": False,
    },
    "IMRD": {
        "opens": None,
        "scholarship": "2027-01-15",
        "starts": "2027-09-01",
        "status": "expected",
        "scholarships": None,
    },
}

COUNTRY_CODES = {
    "NL": "Netherlands",
    "SE": "Sweden",
    "EE": "Estonia",
    "BE": "Belgium",
    "AT": "Austria",
    "FR": "France",
    "CZ": "Czechia",
    "DE": "Germany",
    "ES": "Spain",
    "PT": "Portugal",
    "SI": "Slovenia",
    "UK": "United Kingdom",
    "FI": "Finland",
    "DK": "Denmark",
    "IT": "Italy",
    "GR": "Greece",
    "PL": "Poland",
}


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for env_file in ENV_FILES:
        if not env_file.exists():
            continue
        for line in env_file.read_text(encoding="utf-8").splitlines():
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            values[key.strip()] = val.strip().strip('"').strip("'")
    values.update({k: v for k, v in os.environ.items() if v})
    return values


def slugify(acronym: str) -> str:
    s = acronym.strip().lower()
    s = s.replace("+", "-plus")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def countries_from(text: str) -> list[str]:
    found: list[str] = []
    for code, name in COUNTRY_CODES.items():
        if re.search(rf"\({code}\)", text):
            if name not in found:
                found.append(name)
    if "Netherlands" in text and "Netherlands" not in found:
        found.append("Netherlands")
    if "Poland" in text and "Poland" not in found:
        found.append("Poland")
    if "Spain" in text and "Spain" not in found:
        found.append("Spain")
    return found


def date_status_from(text: str) -> str:
    t = text.upper()
    if "WATCH" in t:
        return "tba"
    if "TBA" in t or "UNCLEAR" in t:
        return "tba"
    if "CONFIRMED" in t and "EXPECTED" not in t:
        return "confirmed"
    if "EXPECTED" in t or "CONFIRMED" in t:
        return "expected"
    return "expected"


def scholarships_flag(text: str) -> bool | None:
    t = text.lower()
    if t.startswith("not ") or "not for 2026" in t or "no erasmus" in t:
        return False
    if "unclear" in t or t.startswith("historically") or t.startswith("likely"):
        return None
    if t.startswith("yes"):
        return True
    return None


class Supabase:
    def __init__(self, url: str, key: str) -> None:
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def request(self, method: str, path: str, payload=None, params: dict | None = None):
        qs = f"?{urllib.parse.urlencode(params, doseq=True)}" if params else ""
        req = urllib.request.Request(
            self.base + path + qs,
            data=None if payload is None else json.dumps(payload).encode(),
            method=method,
            headers={
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "Content-Type": "application/json",
                "Prefer": "return=representation",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as res:
                raw = res.read().decode()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as exc:
            body = exc.read().decode()
            raise SystemExit(f"{method} {path} failed {exc.code}: {body}") from exc

    def get(self, path: str, params: dict) -> list:
        return self.request("GET", path, params=params) or []

    def post(self, path: str, payload: dict) -> dict:
        rows = self.request("POST", path, payload)
        return rows[0]

    def patch(self, path: str, params: dict, payload: dict) -> dict:
        rows = self.request("PATCH", path, payload, params)
        return rows[0]


def read_rows() -> list[dict]:
    wb = load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb["02_Full_catalogue"]
    headers: list[str] = []
    out: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 3:
            continue
        if i == 3:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not row or not row[2]:
            continue
        rec = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
        out.append(rec)
    wb.close()
    return out


def main() -> None:
    env = load_env()
    url = env.get("VITE_SUPABASE_URL") or env.get("SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or env.get("SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("Need VITE_SUPABASE_URL in .env.local and SUPABASE_SERVICE_ROLE_KEY in the environment.")

    db = Supabase(url, key)
    cats = db.get(
        "/catalogues",
        {"slug": "eq.erasmus-mundus", "owner_user_id": "is.null", "select": "id,slug,name"},
    )
    if not cats:
        sys.exit("Erasmus Mundus catalogue not found. Run 001_schema.sql first.")
    catalogue_id = cats[0]["id"]

    existing = {
        row["slug"]: row
        for row in db.get(
            "/programmes",
            {"catalogue_id": f"eq.{catalogue_id}", "select": "id,slug,acronym"},
        )
        if row.get("slug")
    }

    print(f"Catalogue {catalogue_id}", flush=True)
    print(f"Reading {EXCEL.name}…", flush=True)
    rows = read_rows()
    print(f"{len(rows)} programmes in Excel", flush=True)

    imported = 0
    for rec in rows:
        acronym = str(rec["Acronym"]).strip()
        slug = slugify(acronym)
        dates = DATES.get(acronym, {})
        coordinator = str(rec.get("Coordinator / consortium") or "")
        website = str(rec.get("Official website") or "") or None
        date_status = dates.get("status") or date_status_from(str(rec.get("Date status") or ""))
        scholarships = dates.get("scholarships")
        if scholarships is None:
            scholarships = scholarships_flag(str(rec.get("EMJM scholarships for 2027 intake?") or ""))

        payload = {
            "catalogue_id": catalogue_id,
            "slug": slug,
            "acronym": acronym,
            "name": rec.get("Official programme title") or acronym,
            "official_title": rec.get("Official programme title"),
            "coordinator": coordinator or None,
            "consortium": coordinator or None,
            "countries": countries_from(coordinator),
            "duration": rec.get("Duration"),
            "language": rec.get("Language") or "English",
            "website": website,
            "apply_url": website,
            "degree_level": "masters",
            "priority": rec.get("Priority"),
            "fit_notes": rec.get("Why it matches your BS GIS"),
            "extra": {
                "apply": rec.get("Apply?"),
                "target_start": rec.get("Target start"),
                "opens_text": rec.get("Application opens"),
                "deadline_text": rec.get("Scholarship deadline (you = Partner Country / Pakistan)"),
                "self_funded_text": rec.get("Self-funded / later deadline"),
                "scholarships_text": rec.get("EMJM scholarships for 2027 intake?"),
                "prepare": rec.get("What to prepare"),
                "watch_outs": rec.get("Notes / watch-outs"),
                "last_checked": rec.get("Last checked"),
                "source": "Erasmus_Mundus_GIS_Catalogue.xlsx",
            },
        }

        if slug in existing:
            prog = db.patch("/programmes", {"id": f"eq.{existing[slug]['id']}"}, payload)
        else:
            prog = db.post("/programmes", payload)
            existing[slug] = prog
        programme_id = prog["id"]

        cycles = db.get(
            "/programme_cycles",
            {
                "programme_id": f"eq.{programme_id}",
                "label": f"eq.{CYCLE_LABEL}",
                "select": "id",
            },
        )
        cycle_payload = {
            "programme_id": programme_id,
            "label": CYCLE_LABEL,
            "starts_on": dates.get("starts"),
            "application_opens_on": dates.get("opens"),
            "scholarship_deadline": dates.get("scholarship"),
            "date_status": date_status,
            "scholarships_available": scholarships,
            "extra_dates": {
                "opens_text": rec.get("Application opens"),
                "deadline_text": rec.get("Scholarship deadline (you = Partner Country / Pakistan)"),
                "self_funded_text": rec.get("Self-funded / later deadline"),
                "target_start": rec.get("Target start"),
            },
        }
        if cycles:
            db.patch("/programme_cycles", {"id": f"eq.{cycles[0]['id']}"}, cycle_payload)
        else:
            db.post("/programme_cycles", cycle_payload)
        imported += 1
        print(f"  {acronym:12} {date_status:10} {dates.get('scholarship') or 'no deadline yet'}", flush=True)

    print(f"Imported {imported} Erasmus Mundus programmes into catalogue {catalogue_id}")


if __name__ == "__main__":
    main()
